import os
import re
import calendar
import pandas as pd
from openpyxl import load_workbook

RAW_DIR = "data/raw"
OUT_FILE = "data/processed/master_dataset.csv"
AUDIT_FILE = "data/processed/audit_totals.csv"

MONTHS = {
    "gennaio": 1, "febbraio": 2, "marzo": 3, "aprile": 4,
    "maggio": 5, "giugno": 6, "luglio": 7, "agosto": 8,
    "settembre": 9, "ottobre": 10, "novembre": 11, "dicembre": 12,
}

CHANNEL_GROUPS = {
    "POS": "digital",
    "TICKET": "digital",
    "SATISPAY": "digital",
    "CONTANTI": "cash",
    "UBER": "delivery",
    "JUST EAT": "delivery",
    "GLOVO": "delivery",
    "DELIVEROO": "delivery",
}

RAW_CHANNELS = list(CHANNEL_GROUPS.keys())


def extract_year(filename):
    match = re.search(r"20\d{2}", filename)
    if not match:
        raise ValueError(f"Anno non trovato nel nome file: {filename}")
    return int(match.group())


def detect_month(sheet_name):
    s = sheet_name.lower().strip()
    for name, num in MONTHS.items():
        if name in s:
            return num
    return None


def clean_money(value):
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)

    value = str(value).strip()
    if value == "":
        return 0.0

    value = (
        value.replace("€", "")
        .replace(" ", "")
        .replace("\u00a0", "")
        .replace(".", "")
        .replace(",", ".")
    )

    try:
        return float(value)
    except ValueError:
        return 0.0


def normalize_text(value):
    if value is None:
        return ""
    return str(value).strip().upper()


def find_days_row(ws):
    for row in range(1, min(ws.max_row, 30) + 1):
        values = [
            normalize_text(ws.cell(row=row, column=col).value)
            for col in range(1, min(ws.max_column, 40) + 1)
        ]
        if any(v == "GIORNI" for v in values):
            return row
    return None


def get_day_columns(ws, days_row, year, month):
    day_cols = []
    max_day = calendar.monthrange(year, month)[1]

    for col in range(1, ws.max_column + 1):
        value = ws.cell(row=days_row, column=col).value

        if value is None:
            continue

        text_value = str(value).strip()

        if text_value.upper() == "TOT":
            continue

        try:
            day = int(float(text_value))
        except ValueError:
            continue

        if 1 <= day <= max_day:
            day_cols.append((col, day))

    return day_cols


def find_channel_rows(ws):
    found = {}

    for row in range(1, ws.max_row + 1):
        row_text = " ".join(
            normalize_text(ws.cell(row=row, column=col).value)
            for col in range(1, min(ws.max_column, 8) + 1)
        )

        for raw_channel in RAW_CHANNELS:
            if raw_channel in row_text and raw_channel not in found:
                found[raw_channel] = row

        if "TOTALE INCASSO" in row_text and "TOTALE INCASSO" not in found:
            found["TOTALE INCASSO"] = row

    return found


def parse_sheet(ws, year, month, source_file, sheet_name):
    rows = []
    audit_rows = []

    days_row = find_days_row(ws)
    if days_row is None:
        return rows, audit_rows

    day_cols = get_day_columns(ws, days_row, year, month)
    channel_rows = find_channel_rows(ws)

    for col, day in day_cols:
        date = pd.Timestamp(year=year, month=month, day=day)

        channel_values = {}

        for raw_channel in RAW_CHANNELS:
            row = channel_rows.get(raw_channel)
            value = 0.0

            if row is not None:
                value = clean_money(ws.cell(row=row, column=col).value)

            channel_values[raw_channel] = value

        digital = (
            channel_values["POS"]
            + channel_values["TICKET"]
            + channel_values["SATISPAY"]
        )

        delivery = (
            channel_values["UBER"]
            + channel_values["JUST EAT"]
            + channel_values["GLOVO"]
            + channel_values["DELIVEROO"]
        )

        cash = channel_values["CONTANTI"]
        computed_total = digital + delivery + cash

        excel_total = 0.0
        total_row = channel_rows.get("TOTALE INCASSO")
        if total_row is not None:
            excel_total = clean_money(ws.cell(row=total_row, column=col).value)

        rows.append({
            "date": date,
            "year": year,
            "month": month,
            "day": day,
            "digital": digital,
            "delivery": delivery,
            "cash": cash,
            "total": computed_total,
            "pos": channel_values["POS"],
            "ticket": channel_values["TICKET"],
            "satispay": channel_values["SATISPAY"],
            "uber": channel_values["UBER"],
            "just_eat": channel_values["JUST EAT"],
            "glovo": channel_values["GLOVO"],
            "deliveroo": channel_values["DELIVEROO"],
            "source_file": source_file,
            "sheet": sheet_name,
        })

        audit_rows.append({
            "date": date,
            "excel_total": excel_total,
            "computed_total": computed_total,
            "difference": computed_total - excel_total,
            "source_file": source_file,
            "sheet": sheet_name,
        })

    return rows, audit_rows


def main():
    all_rows = []
    all_audit = []

    files = [
        f for f in os.listdir(RAW_DIR)
        if f.lower().endswith(".xlsx") and not f.startswith("~$")
    ]

    if not files:
        raise FileNotFoundError(f"Nessun file Excel trovato in {RAW_DIR}")

    for filename in sorted(files):
        path = os.path.join(RAW_DIR, filename)
        year = extract_year(filename)

        wb = load_workbook(path, data_only=True)

        for sheet_name in wb.sheetnames:
            month = detect_month(sheet_name)
            if month is None:
                continue

            ws = wb[sheet_name]

            rows, audit = parse_sheet(
                ws=ws,
                year=year,
                month=month,
                source_file=filename,
                sheet_name=sheet_name,
            )

            all_rows.extend(rows)
            all_audit.extend(audit)

    df = pd.DataFrame(all_rows)
    audit_df = pd.DataFrame(all_audit)

    if df.empty:
        raise ValueError("Dataset vuoto: non sono stati estratti dati dagli Excel.")

    df = df.sort_values("date")
    audit_df = audit_df.sort_values("date")

    df["delivery_share"] = df["delivery"] / df["total"].replace(0, pd.NA)
    df["digital_share"] = df["digital"] / df["total"].replace(0, pd.NA)
    df["cash_share"] = df["cash"] / df["total"].replace(0, pd.NA)

    os.makedirs("data/processed", exist_ok=True)

    df.to_csv(OUT_FILE, index=False)
    audit_df.to_csv(AUDIT_FILE, index=False)

    print("\n=== MASTER DATASET CREATO ===")
    print(df.head(15))
    print("\nROWS:", len(df))
    print("DATE RANGE:", df["date"].min(), "→", df["date"].max())

    print("\n=== MEDIA CANALI ===")
    print("Delivery share:", round(df["delivery_share"].mean(), 4))
    print("Digital share:", round(df["digital_share"].mean(), 4))
    print("Cash share:", round(df["cash_share"].mean(), 4))

    print("\n=== AUDIT TOTALI ===")
    print(audit_df[["date", "excel_total", "computed_total", "difference"]].head(20))
    print("\nDifferenza media:", round(audit_df["difference"].mean(), 2))
    print("Differenza max assoluta:", round(audit_df["difference"].abs().max(), 2))

    print(f"\nSalvato: {OUT_FILE}")
    print(f"Audit salvato: {AUDIT_FILE}")


if __name__ == "__main__":
    main()

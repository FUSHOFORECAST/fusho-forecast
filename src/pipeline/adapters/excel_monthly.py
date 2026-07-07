import calendar
import fnmatch
import os
import re

import pandas as pd
from openpyxl import load_workbook

from src.pipeline.adapters.common import add_channel_shares, clean_money, normalize_text, slugify
from src.pipeline.config import RestaurantConfig

MONTHS_BY_LANGUAGE = {
    "it": {
        "gennaio": 1, "febbraio": 2, "marzo": 3, "aprile": 4,
        "maggio": 5, "giugno": 6, "luglio": 7, "agosto": 8,
        "settembre": 9, "ottobre": 10, "novembre": 11, "dicembre": 12,
    },
}


def extract_year(filename: str, year_regex: str) -> int:
    match = re.search(year_regex, filename)
    if not match:
        raise ValueError(f"Anno non trovato nel nome file: {filename}")
    return int(match.group())


def detect_month(sheet_name: str, months: dict[str, int]) -> int | None:
    s = sheet_name.lower().strip()
    for name, num in months.items():
        if name in s:
            return num
    return None


def find_days_row(ws, max_scan_rows: int = 30, max_scan_cols: int = 40) -> int | None:
    for row in range(1, min(ws.max_row, max_scan_rows) + 1):
        values = [
            normalize_text(ws.cell(row=row, column=col).value)
            for col in range(1, min(ws.max_column, max_scan_cols) + 1)
        ]
        if any(v == "GIORNI" for v in values):
            return row
    return None


def get_day_columns(ws, days_row: int, year: int, month: int) -> list[tuple[int, int]]:
    day_cols = []
    max_day = calendar.monthrange(year, month)[1]

    for col in range(1, ws.max_column + 1):
        value = ws.cell(row=days_row, column=col).value

        if value is None:
            continue

        text_value = str(value).strip()

        if text_value.upper() == "TOT":
            continue

        try:
            day = int(float(text_value))
        except ValueError:
            continue

        if 1 <= day <= max_day:
            day_cols.append((col, day))

    return day_cols


def find_channel_rows(ws, raw_channels: list[str], total_row_label: str, scan_cols: int = 8) -> dict[str, int]:
    found = {}

    for row in range(1, ws.max_row + 1):
        row_text = " ".join(
            normalize_text(ws.cell(row=row, column=col).value)
            for col in range(1, min(ws.max_column, scan_cols) + 1)
        )

        for raw_channel in raw_channels:
            if raw_channel.upper() in row_text and raw_channel not in found:
                found[raw_channel] = row

        if total_row_label.upper() in row_text and total_row_label not in found:
            found[total_row_label] = row

    return found


def parse_sheet(ws, year: int, month: int, source_file: str, sheet_name: str, config: RestaurantConfig):
    rows = []
    audit_rows = []

    days_row = find_days_row(ws)
    if days_row is None:
        return rows, audit_rows

    day_cols = get_day_columns(ws, days_row, year, month)

    raw_channels = list(config.channel_map.keys())
    channel_rows = find_channel_rows(ws, raw_channels, config.audit.total_row_label)

    for col, day in day_cols:
        date = pd.Timestamp(year=year, month=month, day=day)

        raw_values = {}
        for raw_channel in raw_channels:
            row = channel_rows.get(raw_channel)
            value = 0.0
            if row is not None:
                value = clean_money(ws.cell(row=row, column=col).value)
            raw_values[raw_channel] = value

        channel_totals = {channel: 0.0 for channel in config.channels}
        for raw_channel, value in raw_values.items():
            channel = config.channel_map[raw_channel]
            channel_totals[channel] += value

        computed_total = sum(channel_totals.values())

        declared_total = 0.0
        total_row = channel_rows.get(config.audit.total_row_label)
        if total_row is not None:
            declared_total = clean_money(ws.cell(row=total_row, column=col).value)

        row_dict = {
            "date": date,
            "year": year,
            "month": month,
            "day": day,
            "total": computed_total,
            "source_file": source_file,
            "sheet": sheet_name,
        }
        row_dict.update(channel_totals)
        row_dict.update({slugify(k): v for k, v in raw_values.items()})

        rows.append(row_dict)

        audit_rows.append({
            "date": date,
            "declared_total": declared_total,
            "computed_total": computed_total,
            "difference": computed_total - declared_total,
            "source_file": source_file,
            "sheet": sheet_name,
        })

    return rows, audit_rows


def extract(config: RestaurantConfig, persist: bool = True) -> tuple[pd.DataFrame, pd.DataFrame]:
    months = MONTHS_BY_LANGUAGE.get(config.language)
    if months is None:
        raise ValueError(f"Lingua non supportata per il rilevamento mesi: {config.language}")

    raw_dir = config.raw_dir
    files = [
        f for f in os.listdir(raw_dir)
        if fnmatch.fnmatch(f, config.data_source.file_pattern) and not f.startswith("~$")
    ]

    if not files:
        raise FileNotFoundError(f"Nessun file trovato in {raw_dir} (pattern: {config.data_source.file_pattern})")

    all_rows = []
    all_audit = []

    for filename in sorted(files):
        path = raw_dir / filename
        year = extract_year(filename, config.data_source.year_regex)

        wb = load_workbook(path, data_only=True)

        for sheet_name in wb.sheetnames:
            month = detect_month(sheet_name, months)
            if month is None:
                continue

            ws = wb[sheet_name]

            rows, audit = parse_sheet(
                ws=ws, year=year, month=month,
                source_file=filename, sheet_name=sheet_name, config=config,
            )

            all_rows.extend(rows)
            all_audit.extend(audit)

    df = pd.DataFrame(all_rows)
    audit_df = pd.DataFrame(all_audit)

    if df.empty:
        raise ValueError("Dataset vuoto: non sono stati estratti dati dagli Excel.")

    df = df.sort_values("date").reset_index(drop=True)
    audit_df = audit_df.sort_values("date").reset_index(drop=True)

    df = add_channel_shares(df, config.channels)

    if persist:
        os.makedirs(config.processed_dir, exist_ok=True)
        df.to_csv(config.processed_path("master_dataset.csv"), index=False)
        audit_df.to_csv(config.processed_path("audit_totals.csv"), index=False)

    return df, audit_df
